"""Agent 的工具。

**终结工具**（M10）——返回的就是给用户的最终答案，Agent 不复述、不加工：
    answer_kb          回答 ERP 问题。内部 = 直路整条，正文流式直通前端
    whoami             自我介绍。和寒暄短路共用同一份文本，不许各写一份

**普通工具**——返回的是给 Agent 看的材料，由它组织成话：
    current_time       当前日期时间（北京时间）
    my_documents       用户自己上传了哪些文档（**只列自己的**）
    search_kb          检索知识库（**自动带当前用户的 owner 过滤**）
                       ⚠️ M10 起**不再挂在主 Agent 上**，见 agent.py 的 TOOLS
    save_requirement   记一条需求，并告诉模型还缺什么
    generate_plan      用 Pydantic 约束的结构化输出生成配置清单
    export_excel       落成 xlsx 供下载

⭐ **这个二分是 M10 的全部要点。** M7 的 Agent 只有普通工具，于是 ERP 答案是
Agent 拿着原始材料自己写的——41 题上准确率 87.8%、幻觉率 12.5%，四条成因
（检索词漂移 / 材料串味 / 拒答闸门变软 / 换了 prompt）全都源于「它有那支笔」。
M10 不是把 Agent 调得更听话，是**把笔拿走**。

三条贯穿的规矩：

1. **工具的入参里绝不出现 user_id。** 它从 `ctx.deps` 来，只可能是 cookie 里那个人。
   让它变成入参 = 一句 prompt injection 就能读别人的私有文档。
2. **工具失败要返回一句人话，不要抛异常。** 抛出去会让整轮对话崩掉；
   返回「检索失败，请换个说法」的话，Agent 还能继续。这就是 M7
   「单工具报错不能整轮崩」的落地方式。
3. **工具返回给模型的文本要短。** 它每一个字都会进下一轮的 prompt。
   检索结果给正文，不给整块 JSON；清单生成只回摘要，明细留在库里。
"""

from __future__ import annotations

import logging
import re
import uuid
from datetime import UTC, datetime, timedelta, timezone

from pydantic_ai import RunContext

from copilot.agent.checklist import REQUIREMENT_FIELDS, Checklist
from copilot.agent.deps import AgentDeps
from copilot.config import get_settings
from copilot.retrieve import search

logger = logging.getLogger(__name__)

# 正文里的图片标记 `[图3]`——build_context() 已经重编过号
_IMG_RE = re.compile(r"\[图(\d+)\]")


async def answer_kb_for_deps(deps: AgentDeps) -> str:
    """运行终结知识库工具的实际实现。

    主 Agent 正常通过 ``answer_kb`` 调用它；runner 的硬防线也会在模型漏调工具时
    调用它。两条入口共用同一份实现，避免回退路径悄悄长成第二套 RAG。
    """
    # ⚠️ 一轮只允许一次终结答案。第二次调用会毁掉第一次的引用编号——
    # 那批 [1][2] 已经连着正文流给用户了，而 `citations` 只有一份
    if deps.final_answer is not None:
        return "这一轮已经回答过用户了。不要再调用，直接结束。"
    if deps.llm is None:  # 只可能是接线漏了，不是用户输入能导致的
        logger.error("answer_kb 没拿到 llm，deps 接线漏了")
        return "回答功能暂时不可用。"

    from starlette.concurrency import iterate_in_threadpool

    from copilot.qa import ask_stream

    buf: list[str] = []
    try:
        streamed = await ask_stream(
            deps.session,
            deps.question,
            deps.embedder,
            deps.reranker,
            deps.llm,
            # ⚠️ 隔离红线：同 search_kb，owner 过滤只能来自 deps
            user_id=deps.user_id,
            history=deps.history,
            mode=deps.mode,
            general=deps.general,
        )
        # 配图在正文之前发，理由见 deps.emit_images()
        deps.images = list(streamed.images)
        await deps.emit_images()

        async for kind, piece in iterate_in_threadpool(streamed.stream):
            # Agent 这一路只要正文。推理草稿（详解档才有）先丢掉——
            # Agent 自己的过程展示走的是 tool trace，再叠一层草稿只会更吵
            if kind != "content":
                continue
            buf.append(piece)
            await deps.emit_text(piece)
    except Exception as e:  # noqa: BLE001 - 见文件头第 2 条
        logger.warning("answer_kb 失败：%s", e, exc_info=True)
        if not buf:
            return "查知识库的时候出错了（外部服务异常）。可以让我再试一次。"
        # 已经吐出去一半了，收不回来。把它定成本轮答案，别让 Agent 再写一段
        deps.final_answer = "".join(buf)
        return "回答到一半出错了，已生成的部分用户已经看到。告诉他可以重试。"

    deps.final_answer = "".join(buf)
    # ⚠️ **覆盖而不是合并。** 用户看到的 [1][2] 来自直路自己的编号；
    # `deps.citations` 里原有的是普通工具留下的（那些材料用户根本没看到），
    # 合并会让编号和正文对不上——点开 [1] 看到的是另一篇文档。
    deps.citations = [c.to_dict() for c in streamed.citations]
    deps.private_hits = max(deps.private_hits, streamed.private_hits)
    deps.retrieved.append(streamed.context_text)
    return "已经把答案直接给用户了。不要复述、不要补充、不要总结，本轮到此结束。"


async def answer_kb(ctx: RunContext[AgentDeps]) -> str:
    """回答用户关于旺店通旗舰版 ERP 的问题：操作步骤、参数配置、异常排查、
    功能限制、界面路径——**只要沾边就调它**。答案会直接呈现给用户，
    你不需要、也不允许再复述或补充。

    这个工具没有参数：它用的就是用户这一轮的原话。
    """
    return await answer_kb_for_deps(ctx.deps)


async def whoami(ctx: RunContext[AgentDeps]) -> str:
    """介绍你自己：你是什么、能做什么、不能做什么。
    用户问「你是谁」「你能干什么」「怎么用」时调它。

    这段介绍会直接呈现给用户，你不需要、也不允许再复述或补充。
    """
    from copilot.qa import canned_reply

    deps = ctx.deps
    if deps.final_answer is not None:
        return "这一轮已经回答过用户了。不要再调用，直接结束。"

    # ⚠️ **不要自己写这段介绍。** 它和寒暄短路那条路必须是同一份文本
    # （见 qa.canned_reply）——两处各写一份，加了新能力就会有一处漏改，
    # 而用户看到的是「同一个助手对自己的两种说法」
    text = canned_reply("capability") or ""
    await deps.emit_text(text)
    deps.final_answer = text
    return "已经把介绍直接给用户了。不要复述、不要补充，本轮到此结束。"


async def my_documents(ctx: RunContext[AgentDeps]) -> str:
    """列出用户自己上传到私有知识库的文档。
    用户问「我传了哪些文档」「我的知识库里有什么」时调它。
    """
    from sqlalchemy import desc, select

    from copilot.db.models import Document

    deps = ctx.deps
    stmt = (
        select(Document.title, Document.status)
        # ⚠️ 隔离红线：同 answer_kb，owner 只能来自 deps。
        # 这里是**严格等于**当前用户，公共库那 746 篇（owner_id IS NULL）
        # 不属于任何人，不该出现在「我的文档」里
        .where(Document.owner_id == deps.user_id)
        .order_by(desc(Document.created_at))
        .limit(20)
    )
    try:
        rows = list((await deps.session.execute(stmt)).all())
    except Exception as e:  # noqa: BLE001 - 见文件头第 2 条
        logger.warning("my_documents 失败：%s", e, exc_info=True)
        return "查文档列表时出错了。可以让我再试一次。"

    if not rows:
        return "这个用户的私有知识库是空的，还没有上传过文档。"
    lines = [
        f"- {title}" + ("" if status == "done" else f"（{status}，还不能被检索到）")
        for title, status in rows
    ]
    more = "（只列最近 20 份）" if len(rows) == 20 else ""
    return f"用户自己上传了 {len(rows)} 份文档{more}：\n" + "\n".join(lines)


# 北京时间。用固定 +8 而不是 ZoneInfo("Asia/Shanghai")：中国 1991 年起
# 全境单一时区、无夏令时，固定偏移是准确的；而 ZoneInfo 在 Windows 上要额外
# 装 tzdata，为了一个不会变的偏移量多一个依赖不值得。
_CST = timezone(timedelta(hours=8))
_WEEKDAYS = ("一", "二", "三", "四", "五", "六", "日")


def current_time() -> str:
    """当前的日期和时间（北京时间）。用户问「现在几点」「今天几号」「星期几」时调它。"""
    now = datetime.now(_CST)
    return f"{now:%Y年%m月%d日 %H:%M}，星期{_WEEKDAYS[now.weekday()]}（北京时间）"


async def search_kb(ctx: RunContext[AgentDeps], query: str) -> str:
    """检索旺店通旗舰版 ERP 知识库，返回带编号来源的原文片段。

    Args:
        query: 检索用的问题或关键词。用中文，尽量具体（含界面名、功能名更准）。
    """
    deps = ctx.deps
    try:
        result = await search(
            deps.session,
            query,
            deps.embedder,
            deps.reranker,
            # ⚠️ 隔离红线：owner 过滤只能来自 deps，不能是入参
            user_id=deps.user_id,
        )
    except Exception as e:  # noqa: BLE001 - 见文件头第 2 条
        logger.warning("search_kb 失败 query=%r：%s", query[:60], e, exc_info=True)
        return "检索失败了（外部服务异常）。可以换个说法再试一次，或者先问别的。"

    if result.is_empty:
        return f"知识库里没有检索到与「{query}」相关的内容。"

    bundle = result.build_context()
    deps.private_hits = max(deps.private_hits, result.private_count)
    renumbered = deps.merge_citations([c.to_dict() for c in result.citations])

    # 把本次检索的 [n] 换成本轮全局编号，否则一轮里检索两次会出现两个 [1]
    mapping = {old["n"]: new["n"] for old, new in zip(
        [c.to_dict() for c in result.citations], renumbered, strict=True
    )}
    text = bundle.text
    for old_n in sorted(mapping, reverse=True):  # 从大到小替换，避免 [1]→[10] 又被再替一次
        text = text.replace(f"[{old_n}] 来源：", f"[{mapping[old_n]}] 来源：")

    # 配图也要并进本轮总表，并把正文里的 [图N] 改成全局号
    if bundle.images:
        offset = len(deps.images)
        for img in bundle.images:
            deps.images.append({"n": img["n"] + offset, "url": img["url"]})
        if offset:
            text = _IMG_RE.sub(lambda m: f"[图{int(m.group(1)) + offset}]", text)

    deps.retrieved.append(text)
    return text


async def save_requirement(ctx: RunContext[AgentDeps], field: str, value: str) -> str:
    """记录一条客户的实施需求，返回还缺哪些信息。

    Args:
        field: 需求字段名，只能是这几个之一：
            platforms（对接平台）、shop_count（店铺数量）、
            warehouse_mode（仓库模式）、warehouse_count（仓库数量）、
            daily_orders（日均单量）、logistics（物流商）、specials（特殊业务）
        value: 客户说的原话或归纳后的值，中文。
    """
    if field not in REQUIREMENT_FIELDS:
        # 不抛异常：告诉模型正确的取值范围，它下一步就能改对（比整轮崩掉好得多）
        return f"没有 `{field}` 这个字段。可用字段：{'、'.join(REQUIREMENT_FIELDS)}"

    value = (value or "").strip()
    if not value:
        return f"`{field}` 的值是空的，没有记录。"

    current = getattr(ctx.deps.profile, field)
    correction_markers = (
        "说错",
        "改成",
        "更正",
        "不是",
        "应为",
        "应该是",
        "调整为",
        "更新为",
    )
    explicit_correction = any(marker in ctx.deps.question for marker in correction_markers)
    if current and current != value and not explicit_correction:
        # 模型有时会把「自营，一个仓」里的“自营”误写到 platforms，覆盖前面
        # 已确认的「淘宝、拼多多」。没有明确更正语气时，已填字段只读；让模型
        # 转去记录真正缺失的字段，避免一轮误判污染整条会话。
        label = REQUIREMENT_FIELDS[field][0]
        ctx.deps.noop_saves += 1
        # ⚠️ **第二次空转就必须叫停。** 2026-08-23 线上：一条收集过需求的会话里，
        # 用户问「星辰电商的对账以什么为准」，模型把这句话当成需求，一个字段
        # 一个字段地试着记，每次都被这里挡回去、每次都换一个字段再试，
        # 一路撞穿 `tool_calls_limit`，整轮炸掉、一个字都没答。
        # 而挡回去时给的话是「请记录其他缺失字段」——那正是让它继续试的指令。
        if ctx.deps.noop_saves >= 2:
            return (
                f"`{field}`（{label}）已记录为「{current}」，未覆盖。"
                "⚠️ 本轮**不要再调用 save_requirement**：用户这一句不是在补充需求。"
                "直接回答用户的问题；如果是知识库问题，调用 answer_kb。"
            )
        return (
            f"`{field}`（{label}）已记录为「{current}」，用户本轮没有明确要求修改，"
            "因此未覆盖。请记录本轮对应的其他缺失字段。"
        )

    setattr(ctx.deps.profile, field, value[:200])
    label = REQUIREMENT_FIELDS[field][0]
    gaps = ctx.deps.profile.missing()
    if not gaps:
        return f"已记录「{label}：{value}」。需求已收集齐，可以调用 generate_plan 生成方案了。"
    nxt = REQUIREMENT_FIELDS[gaps[0]]
    return (
        f"已记录「{label}：{value}」。\n"
        f"还缺 {len(gaps)} 项：{'、'.join(REQUIREMENT_FIELDS[f][0] for f in gaps)}。\n"
        f"接下来问这个：{nxt[1]}"
    )


async def generate_plan(ctx: RunContext[AgentDeps]) -> str:
    """根据已收集的需求生成《实施配置方案》。需求收集齐之后再调。"""
    deps = ctx.deps
    # ⚠️ **已经有方案了就别再生成一遍。** 2026-08-23 人工验收：方案出完之后
    # 用户说「你好」「好的谢谢」，模型都照样调了这个工具，于是一句招呼要跑一次
    # 子 Agent 加二十来次检索，用户收到的是**又一整段方案摘要**——而且每次
    # 项数还不一样（14 → 15 → 13），看起来像系统在自说自话地改方案。
    # 真要改需求时，用户会先说改什么，`save_requirement` 更新完 profile
    # 之后由 runner 的收尾逻辑重新生成；这里挡的是「无理由重跑」。
    if deps.checklist is not None:
        return (
            f"方案《{deps.checklist.title}》本轮已经生成过了，不要重复生成。"
            "直接回答用户这一句；用户若要修改需求，先用 save_requirement 更新对应字段。"
        )
    gaps = deps.profile.missing()
    if len(gaps) > 2:
        # 关键信息缺太多时硬生成，只会得到一份「按需配置」的废纸
        return (
            "需求还差得多，先别生成。缺：" + "、".join(REQUIREMENT_FIELDS[f][0] for f in gaps)
        )

    # 惰性导入：planner 会建一个子 Agent，而 tools 在 agent.py 里被引用，
    # 顶层互相 import 会成环
    from copilot.agent.planner import build_checklist

    try:
        checklist = await build_checklist(deps)
    except Exception as e:  # noqa: BLE001
        logger.warning("generate_plan 失败：%s", e, exc_info=True)
        return "生成方案时出错了。可以让我再试一次。"

    deps.checklist = checklist
    lines = [f"已生成《{checklist.title}》，共 {len(checklist.items)} 项配置：", ""]
    for it in checklist.items[:8]:
        lines.append(f"- [{it.priority}] {it.area} → {it.item}：{it.value}")
    if len(checklist.items) > 8:
        lines.append(f"- …另有 {len(checklist.items) - 8} 项")
    if checklist.open_questions:
        lines += ["", "需要人工确认：" + "；".join(checklist.open_questions[:3])]
    lines += ["", "接下来可以调用 export_excel 导出成 xlsx。"]
    return "\n".join(lines)


async def export_excel(ctx: RunContext[AgentDeps]) -> str:
    """把已生成的配置方案导出成 xlsx，返回下载地址。"""
    deps = ctx.deps
    if deps.checklist is None:
        return "还没有方案可以导出，先调用 generate_plan。"

    s = get_settings()
    rel = f"{deps.user_id}/{deps.conversation_id}.xlsx"
    path = s.export_path(rel)
    path.parent.mkdir(parents=True, exist_ok=True)

    try:
        _write_xlsx(path, deps.checklist)
    except Exception as e:  # noqa: BLE001
        logger.warning("导出 xlsx 失败：%s", e, exc_info=True)
        return "导出文件时出错了。方案本身已经生成，可以再试一次导出。"

    deps.download_url = f"/api/conversations/{deps.conversation_id}/export"
    # 告诉模型「已经有下载入口了」，但别让它自己编 URL——真正的地址由前端从
    # data-download 片段拿（见 routes/chat.py）
    return "已导出 xlsx，页面上会出现下载按钮。不要在回答里自己拼下载链接。"


def _write_xlsx(path, checklist: Checklist) -> None:
    """写 xlsx。列宽和换行都得设——不设的话打开是一列挤在一起的天书。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "实施配置方案"

    ws["A1"] = checklist.title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = checklist.summary
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 46

    headers = ["优先级", "模块／路径", "配置项", "建议值", "依据"]
    head_fill = PatternFill("solid", fgColor="EEF2FF")
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=name)
        cell.font = Font(bold=True)
        cell.fill = head_fill

    for i, it in enumerate(checklist.items, start=5):
        for col, val in enumerate(
            [it.priority, it.area, it.item, it.value, it.why], start=1
        ):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")

    for col, width in zip("ABCDE", (8, 26, 22, 40, 52), strict=True):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"  # 表头钉住，几十行的清单往下翻还看得见列名

    if checklist.open_questions:
        sheet2 = wb.create_sheet("待确认")
        sheet2["A1"] = "知识库里没有答案、需要人工确认的问题"
        sheet2["A1"].font = Font(bold=True)
        for i, q in enumerate(checklist.open_questions, start=2):
            sheet2.cell(row=i, column=1, value=q).alignment = Alignment(wrap_text=True)
        sheet2.column_dimensions["A"].width = 80

    meta = wb.create_sheet("生成信息")
    meta["A1"] = "生成时间"
    meta["B1"] = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    meta["A2"] = "生成方式"
    meta["B2"] = "旺店通旗舰版 ERP 知识库助手（依据知识库材料生成，请人工复核）"
    meta.column_dimensions["A"].width = 14
    meta.column_dimensions["B"].width = 60

    wb.save(str(path))


def conversation_export_name(conversation_id: uuid.UUID) -> str:
    """下载时给用户看的文件名。"""
    return f"实施配置方案-{str(conversation_id)[:8]}.xlsx"
