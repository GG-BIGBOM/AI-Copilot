"""M7 Agent 的测试。不调 LLM——工具、状态机、路由都能单独验。

守的四件事：

1. **⭐ `search_kb` 必须带 owner 过滤。** 工具的入参里没有 user_id，它只能从
   `deps` 来。这条破了，一句 prompt injection 就能读别人的私有文档。
2. **工具失败要返回一句人话，不能抛异常。** 抛出去整轮对话就崩了。
3. **需求收集的状态机**：记了什么、还缺什么、缺太多时不许生成。
4. **路由**：普通问答不能被 Agent 接管（它在评测上更差），
   但已经在收集需求的会话必须继续留在 Agent 里。
"""

from __future__ import annotations

import uuid
from types import SimpleNamespace

import pytest
from sqlalchemy import delete

from copilot.agent.checklist import Checklist, ChecklistItem, Requirement
from copilot.agent.deps import AgentDeps
from copilot.agent.tools import export_excel, generate_plan, save_requirement, search_kb
from copilot.config import get_settings
from copilot.db.models import Chunk, Conversation, Document, User

DIM = 1024


class FakeEmbedder:
    dim = DIM

    def __init__(self, boom: Exception | None = None) -> None:
        self.boom = boom

    @staticmethod
    def _vec(text: str) -> list[float]:
        v = [0.0] * DIM
        for i, ch in enumerate(text[:64]):
            v[(ord(ch) * 7 + i) % DIM] += 1.0
        norm = sum(x * x for x in v) ** 0.5 or 1.0
        return [x / norm for x in v]

    def embed_documents(self, texts):
        return [self._vec(t) for t in texts]

    def embed_query(self, text):
        if self.boom:
            raise self.boom
        return self._vec(text)


class PassThroughReranker:
    def rerank(self, query, documents, top_k):
        from copilot.providers.base import RerankResult

        return [RerankResult(index=i, score=1.0) for i in range(min(top_k, len(documents)))]


def ctx(deps: AgentDeps):
    """工具只用 `ctx.deps`，所以测试给个最小替身就够，不必造整个 RunContext。"""
    return SimpleNamespace(deps=deps)


@pytest.fixture
async def two_users(maker):
    """两个用户，各有一篇私有文档。"""
    tag = uuid.uuid4().hex[:8]
    emb = FakeEmbedder()
    async with maker() as s:
        alice = User(email=f"ag-a-{tag}@t.local", password_hash="x")
        bob = User(email=f"ag-b-{tag}@t.local", password_hash="x")
        s.add_all([alice, bob])
        await s.flush()
        secrets = {
            alice.id: f"爱丽丝的私有报价是每单三十七元-{tag}",
            bob.id: f"鲍勃的私有合同金额是五十万元-{tag}",
        }
        for uid, body in secrets.items():
            doc = Document(
                owner_id=uid,
                source_type="upload",
                title=f"私有-{uid.hex[:6]}-{tag}",
                content_hash=uuid.uuid4().hex,
                status="done",
                chunk_count=1,
            )
            s.add(doc)
            await s.flush()
            s.add(
                Chunk(
                    document_id=doc.id,
                    owner_id=uid,
                    ordinal=0,
                    content=body,
                    embedding=emb.embed_query(body),
                    title=doc.title,
                )
            )
        await s.commit()
        ids = (alice.id, bob.id)

    yield ids, secrets, tag

    async with maker() as s:
        for uid in ids:
            docs = list(
                (
                    await s.execute(
                        __import__("sqlalchemy").select(Document.id).where(Document.owner_id == uid)
                    )
                ).scalars()
            )
            if docs:
                await s.execute(delete(Chunk).where(Chunk.document_id.in_(docs)))
                await s.execute(delete(Document).where(Document.id.in_(docs)))
        await s.execute(delete(User).where(User.id.in_(ids)))
        await s.commit()


def _deps(session, user_id, **kw) -> AgentDeps:
    # ⚠️ 不传 `space_id` 就是 fail closed（检索一条都不返回）。这一组验的是
    # Agent 工具的隔离红线（谁的文档），空间那根轴在 test_spaces.py 里单独验。
    from conftest import flagship_space_id

    kw.setdefault("space_id", flagship_space_id())
    return AgentDeps(
        session=session,
        user_id=user_id,
        conversation_id=uuid.uuid4(),
        embedder=FakeEmbedder(),
        reranker=PassThroughReranker(),
        **kw,
    )


# ---------- ⭐ 隔离红线 ----------


async def test_answer_kb_never_crosses_users(maker, two_users):
    """⭐ 最不能红的一条。M10 之后 `answer_kb` 是**活的**检索路径，
    这条红线跟着它走：用 B 的原文当问题、以 A 的身份提问，最严苛的情形。
    """
    from chat_helpers import FakeLLM

    from copilot.agent.tools import answer_kb

    (alice_id, bob_id), secrets, _ = two_users
    async with maker() as s:
        # 问题从 deps 来，不是工具入参——见下一条
        deps = _deps(s, alice_id, llm=FakeLLM("知识库暂无此内容。"), question=secrets[bob_id])
        await answer_kb(ctx(deps))
    assert secrets[bob_id] not in deps.context_text, "泄漏了！爱丽丝检索到了鲍勃的材料"


async def test_answer_kb_finds_own_document(maker, two_users):
    """隔离不能矫枉过正——自己的东西自己得搜到。

    没有这一条，上面那条用「什么都检索不到」也能过。
    """
    from chat_helpers import FakeLLM

    from copilot.agent.tools import answer_kb

    (alice_id, _), secrets, _ = two_users
    async with maker() as s:
        deps = _deps(s, alice_id, llm=FakeLLM("好的[1]。"), question=secrets[alice_id])
        await answer_kb(ctx(deps))
    assert secrets[alice_id] in deps.context_text


def test_answer_kb_takes_no_question_argument():
    """⭐ 结构性防线，两层：

    1. 没有 user_id 入参——同 `search_kb`，否则一句 prompt injection 就能读别人的库。
    2. **连问题都不是入参。** 检索用的是 `deps.question`（用户原话）。
       让模型自己写检索词，就是 M7 准确率掉 12 个点的第一条成因：
       它把「京东电子面单怎么配」改写成「电子面单 模板 设置」，召回了得物那篇。
    """
    import inspect

    from copilot.agent.tools import answer_kb

    params = list(inspect.signature(answer_kb).parameters)
    assert params == ["ctx"], f"answer_kb 多了入参：{params}"


async def test_search_kb_never_crosses_users(maker, two_users):
    """⭐ 同一条红线在 M7 那个工具上的版本。

    用 B 的原文当查询、以 A 的身份检索——最严苛的情形。
    """
    (alice_id, bob_id), secrets, _ = two_users
    async with maker() as s:
        deps = _deps(s, alice_id)
        text = await search_kb(ctx(deps), secrets[bob_id])
    assert secrets[bob_id] not in text, f"泄漏了！爱丽丝拿到了鲍勃的内容：{text[:200]}"


async def test_search_kb_finds_own_document(maker, two_users):
    """隔离不能矫枉过正——自己的东西自己得搜到。"""
    (alice_id, _), secrets, _ = two_users
    async with maker() as s:
        deps = _deps(s, alice_id)
        text = await search_kb(ctx(deps), secrets[alice_id])
    assert secrets[alice_id] in text


def test_search_kb_signature_has_no_user_id():
    """⭐ 结构性防线：工具的入参里**不允许**出现 user_id。

    一旦它成了入参，模型就可以（被诱导着）填别人的 id——
    那意味着一句 prompt injection 就能读到别人的私有文档。
    这条断言比上面两条更根本：它防的是「将来有人图方便加了这个参数」。
    """
    import inspect

    params = set(inspect.signature(search_kb).parameters) - {"ctx"}
    assert params == {"query"}, f"search_kb 的入参变了：{params}"


# ---------- 工具失败不能整轮崩 ----------


async def test_search_kb_returns_message_on_failure(maker, two_users):
    """embedding 服务挂了时返回一句人话，不抛异常——否则整轮对话崩掉。"""
    (alice_id, _), _, _ = two_users
    async with maker() as s:
        deps = _deps(s, alice_id)
        deps.embedder = FakeEmbedder(boom=RuntimeError("429 rate limited"))
        text = await search_kb(ctx(deps), "面单怎么设置")
    assert "失败" in text
    assert "再试" in text


async def test_search_kb_says_nothing_found(maker, two_users):
    (alice_id, _), _, tag = two_users
    async with maker() as s:
        deps = _deps(s, alice_id)
        text = await search_kb(ctx(deps), f"完全不存在的东西{uuid.uuid4().hex}")
    # 检索不到时要明确说没有，别返回空串——空串会让模型以为工具坏了
    assert "没有检索到" in text or "[1]" in text


# ---------- 需求收集的状态机 ----------


async def test_save_requirement_reports_what_is_missing(maker, two_users):
    """记完要告诉模型下一个问什么——追问的节奏靠这个带。"""
    (alice_id, _), _, _ = two_users
    async with maker() as s:
        deps = _deps(s, alice_id)
        out = await save_requirement(ctx(deps), "platforms", "淘宝、抖音")
    assert deps.profile.platforms == "淘宝、抖音"
    assert "还缺" in out
    assert "店铺数量" in out  # 下一个字段的名字


async def test_save_requirement_rejects_unknown_field_without_raising(maker, two_users):
    """字段名填错时告诉它可用范围，别抛异常。模型下一步就能改对。"""
    (alice_id, _), _, _ = two_users
    async with maker() as s:
        deps = _deps(s, alice_id)
        out = await save_requirement(ctx(deps), "budget", "十万")
    assert "没有" in out and "platforms" in out


async def test_save_requirement_announces_completion(maker, two_users):
    (alice_id, _), _, _ = two_users
    filled = Requirement(
        platforms="淘宝",
        shop_count="1",
        warehouse_mode="自营",
        warehouse_count="1",
        daily_orders="500",
        logistics="中通",
    )
    async with maker() as s:
        deps = _deps(s, alice_id, profile=filled)
        out = await save_requirement(ctx(deps), "specials", "组合装")
    assert "收集齐" in out
    assert deps.profile.missing() == []


async def test_save_requirement_does_not_silently_overwrite_existing_field(
    maker, two_users
):
    """仓库模式里的“自营”不能污染已经确认的平台。"""
    (alice_id, _), _, _ = two_users
    async with maker() as s:
        deps = _deps(
            s,
            alice_id,
            profile=Requirement(platforms="淘宝、拼多多"),
            question="自营，一个仓",
        )
        out = await save_requirement(ctx(deps), "platforms", "自营")
    assert deps.profile.platforms == "淘宝、拼多多"
    assert "未覆盖" in out


async def test_save_requirement_allows_explicit_correction(maker, two_users):
    """用户明确纠正时仍可更新，不能把 G6 的正常改口一起锁死。"""
    (alice_id, _), _, _ = two_users
    async with maker() as s:
        deps = _deps(
            s,
            alice_id,
            profile=Requirement(shop_count="5 个店"),
            question="刚才说错了，是 8 个",
        )
        await save_requirement(ctx(deps), "shop_count", "8 个店")
    assert deps.profile.shop_count == "8 个店"


async def test_generate_plan_refuses_when_too_little_known(maker, two_users):
    """信息缺太多时硬生成，只会得到一份写满「按需配置」的废纸。"""
    (alice_id, _), _, _ = two_users
    async with maker() as s:
        deps = _deps(s, alice_id, profile=Requirement(platforms="淘宝"))
        out = await generate_plan(ctx(deps))
    assert "先别生成" in out


async def test_export_without_plan_tells_what_to_do_first(maker, two_users):
    (alice_id, _), _, _ = two_users
    async with maker() as s:
        deps = _deps(s, alice_id)
        out = await export_excel(ctx(deps))
    assert "generate_plan" in out


# ---------- xlsx ----------


async def test_export_writes_a_readable_workbook(maker, two_users):
    """导出的文件要真能打开，且列宽/冻结/工作表都在——
    不设这些的话，打开是一列挤在一起的天书。"""
    from openpyxl import load_workbook

    (alice_id, _), _, _ = two_users
    checklist = Checklist(
        title="测试方案",
        summary="两句话说清思路",
        items=[
            ChecklistItem(
                area="设置–基本设置–店铺",
                item="创建店铺并授权",
                value="平台选淘宝，右键店铺授权",
                why="材料 [1] 写了授权步骤",
                priority="必做",
            )
        ],
        open_questions=["日均单量还需确认"],
    )
    async with maker() as s:
        deps = _deps(s, alice_id, checklist=checklist)
        out = await export_excel(ctx(deps))
    assert "已导出" in out
    assert deps.download_url and str(deps.conversation_id) in deps.download_url

    path = get_settings().export_path(f"{alice_id}/{deps.conversation_id}.xlsx")
    try:
        assert path.exists()
        wb = load_workbook(path)
        assert wb.sheetnames == ["实施配置方案", "待确认", "生成信息"]
        ws = wb["实施配置方案"]
        assert [ws.cell(row=4, column=c).value for c in range(1, 6)] == [
            "优先级",
            "模块／路径",
            "配置项",
            "建议值",
            "依据",
        ]
        assert ws["A5"].value == "必做"
        assert ws.freeze_panes == "A5"  # 表头钉住
    finally:
        path.unlink(missing_ok=True)


# ---------- 路由：谁该走 Agent ----------


async def test_plain_question_does_not_use_agent(maker, two_users):
    """⭐ 普通问答**不能**被 Agent 接管。

    依据是数字：M8 的 41 题上 Agent 准确率 87.8% vs 直路 100%、
    幻觉率 12.5% vs 0%。让它接管等于拿已量化的质量去换没量化的。
    """
    from copilot.api.routes.chat import _use_agent

    (alice_id, _), _, _ = two_users
    async with maker() as s:
        user = await s.get(User, alice_id)
        assert await _use_agent(s, user, "京东电子面单模板怎么设置？", None) is False


async def test_plan_request_uses_agent(maker, two_users):
    from copilot.api.routes.chat import _use_agent

    (alice_id, _), _, _ = two_users
    async with maker() as s:
        user = await s.get(User, alice_id)
        assert await _use_agent(s, user, "帮我出一个实施配置方案", None) is True


@pytest.mark.parametrize(
    ("profile", "title", "question", "expected"),
    [
        (None, "普通问答", "帮我出一个实施方案", True),
        (
            None,
            "普通问答",
            "我有淘宝拼多多抖音三个平台，一共 12 个店，没有特殊业务，帮我出方案",
            True,
        ),
        ({}, "帮我出一个实施方案", "淘宝、拼多多", True),
        ({"platforms": "淘宝"}, "普通问答", "一共 5 个店", True),
        ({}, "普通问答", "电子面单怎么设置", False),
    ],
)
def test_plan_tool_limits_follow_real_flow(profile, title, question, expected):
    from copilot.api.routes.chat import _needs_plan_limits

    conv = SimpleNamespace(profile=profile, title=title)
    assert _needs_plan_limits(conv, question) is expected


async def test_generate_plan_refuses_to_rebuild_what_it_already_built(maker, two_users):
    """第二道闸：就算模型还是调了 generate_plan，也不该重跑一遍。

    重跑一次要开一个子 Agent、二十来次检索，用户等半天收到的是一份
    项数还不一样的「新」方案。
    """
    from types import SimpleNamespace

    from copilot.agent.checklist import Checklist
    from copilot.agent.tools import generate_plan

    (alice_id, _), _, _ = two_users
    async with maker() as s:
        deps = _deps(s, alice_id, profile=Requirement(platforms="淘宝"))
        deps.checklist = Checklist(title="实施配置清单", summary="12 店 2 仓", items=[])
        reply = await generate_plan(SimpleNamespace(deps=deps))

    assert "已经生成过" in reply
    assert "不要重复生成" in reply


async def test_small_talk_comes_back_once_the_plan_is_done(maker, two_users):
    """⭐ 「收集中」和「出过方案」是两回事。

    2026-08-23 人工验收：方案已经生成、也导出了，用户说「你好」——原来仍然
    算在收集流程里，寒暄短路被跳过、整句交给 Agent，模型看着一份完整需求
    **又把方案重新生成了一遍**（页面上是「已参考 21 条知识内容」加一整段
    方案摘要，一句招呼要跑一次子 Agent 加二十来次检索），而且每次项数还不同
    （14 → 15 → 13），看起来像系统在自说自话地改方案。
    """
    from copilot.api.routes.chat import _active_plan_flow
    from copilot.db.models import Conversation

    (alice_id, _), _, _ = two_users
    async with maker() as s:
        conv = Conversation(
            user_id=alice_id, title="帮我出一个实施方案", profile={"platforms": "淘宝"}
        )
        s.add(conv)
        await s.commit()
        cid = str(conv.id)

        assert await _active_plan_flow(s, alice_id, cid) is True, "还在收集，寒暄不该短路"

        conv.checklist = {"title": "实施配置清单", "items": []}
        await s.commit()

        assert await _active_plan_flow(s, alice_id, cid) is False, (
            "方案已经出完了，收集流程就结束了，寒暄要恢复短路"
        )


@pytest.mark.parametrize(
    ("profile", "note"),
    [
        ({"platforms": "淘宝"}, "已经记了字段"),
        # ⭐ 这一条是线上实测踩出来的：第一轮 Agent 往往只是提个问题、
        # 一个字段都没记，profile 是空字典。按「profile 有内容」判的话，
        # 第二轮就掉回直路——用户答「淘宝和抖音两个平台」，
        # 回来的是一句「根据参考材料，无法回答」，对话直接散掉。
        ({}, "刚开始，还没记到任何字段"),
    ],
)
async def test_conversation_in_progress_stays_on_agent(maker, two_users, profile, note):
    """⭐ 已经在走 Agent 的会话必须继续走 Agent。

    判据是 `profile is not None`，不是「profile 有内容」。
    """
    from copilot.api.routes.chat import _use_agent

    (alice_id, _), _, _ = two_users
    async with maker() as s:
        conv = Conversation(user_id=alice_id, title="收集中", profile=profile)
        s.add(conv)
        await s.commit()
        user = await s.get(User, alice_id)
        try:
            # 「一共 5 个店」里没有任何意图词，全靠会话状态把它留在 Agent
            assert await _use_agent(s, user, "一共 5 个店", str(conv.id)) is True, note
        finally:
            await s.execute(delete(Conversation).where(Conversation.id == conv.id))
            await s.commit()


async def test_fresh_conversation_does_not_stick_to_agent(maker, two_users):
    """反例：没走过 Agent 的会话（profile 是 NULL）不该被粘住。"""
    from copilot.api.routes.chat import _use_agent

    (alice_id, _), _, _ = two_users
    async with maker() as s:
        conv = Conversation(user_id=alice_id, title="普通问答")
        s.add(conv)
        await s.commit()
        user = await s.get(User, alice_id)
        try:
            assert await _use_agent(s, user, "面单怎么设置", str(conv.id)) is False
        finally:
            await s.execute(delete(Conversation).where(Conversation.id == conv.id))
            await s.commit()


async def test_someone_elses_conversation_does_not_flip_routing(maker, two_users):
    """别人的会话 id 不能影响我的路由（顺手也是一道越权检查）。"""
    from copilot.api.routes.chat import _use_agent

    (alice_id, bob_id), _, _ = two_users
    async with maker() as s:
        conv = Conversation(user_id=bob_id, title="鲍勃在收集", profile={"platforms": "拼多多"})
        s.add(conv)
        await s.commit()
        alice = await s.get(User, alice_id)
        try:
            assert await _use_agent(s, alice, "一共 5 个店", str(conv.id)) is False
        finally:
            await s.execute(delete(Conversation).where(Conversation.id == conv.id))
            await s.commit()


# ---------- 消息历史 ----------


def test_message_history_keeps_order_and_roles():
    """多轮追问全靠它。转错了 Agent 就会失忆、把问过的再问一遍。"""
    from pydantic_ai.messages import ModelRequest, ModelResponse

    from copilot.agent.runner import to_message_history

    msgs = to_message_history(
        [("user", "帮我出方案"), ("assistant", "先问平台"), ("user", "淘宝"), ("tool", "忽略")]
    )
    assert [type(m) for m in msgs] == [ModelRequest, ModelResponse, ModelRequest]
    assert msgs[0].parts[0].content == "帮我出方案"


def test_message_history_skips_empty():
    from copilot.agent.runner import to_message_history

    assert to_message_history([("user", "  "), ("assistant", "")]) == []


# ---------- 引用编号 ----------


def test_citations_are_renumbered_across_searches(maker):
    """Agent 一轮里会检索好几次，每次的引用都从 [1] 开始。

    不重编号的话答案里会出现两个 [1]，用户点开溯源看到的是另一篇文档。
    """
    deps = AgentDeps(
        session=None,  # 这个用例不碰数据库
        user_id=uuid.uuid4(),
        conversation_id=uuid.uuid4(),
        embedder=FakeEmbedder(),
    )
    first = deps.merge_citations([{"n": 1, "title": "甲", "heading": "一"}])
    second = deps.merge_citations(
        [{"n": 1, "title": "乙", "heading": "二"}, {"n": 2, "title": "甲", "heading": "一"}]
    )
    assert [c["n"] for c in first] == [1]
    assert [c["n"] for c in second] == [2, 1]  # 乙拿到新号 2；甲复用已有的 1
    assert len(deps.citations) == 2, "同一篇不该占两个号"


# ---------- M10 P3：灰度分桶 ----------


def _settings(monkeypatch, **kw):
    """改配置单例上的字段，用完自动还原。"""
    s = get_settings()
    for k, v in kw.items():
        monkeypatch.setattr(s, k, v)
    return s


def test_rollout_zero_and_one_are_absolute(monkeypatch):
    from copilot.api.routes.chat import in_agent_bucket

    users = [uuid.uuid4() for _ in range(200)]
    _settings(monkeypatch, agent_enabled=False, agent_rollout=0.0)
    assert not any(in_agent_bucket(u) for u in users)
    _settings(monkeypatch, agent_rollout=1.0)
    assert all(in_agent_bucket(u) for u in users)


def test_rollout_bucket_is_stable_for_the_same_user(monkeypatch):
    """⭐ 分桶必须按用户，不能按请求。

    同一个人在两条路之间跳，多轮收集需求的状态当场断掉；线上出问题时
    也归不了因——你不知道他刚才那一句走的是哪条路。
    """
    from copilot.api.routes.chat import in_agent_bucket

    _settings(monkeypatch, agent_enabled=False, agent_rollout=0.5)
    user = uuid.uuid4()
    assert len({in_agent_bucket(user) for _ in range(50)}) == 1


def test_rollout_ratio_is_roughly_right(monkeypatch):
    """10% 的桶不该装进 30% 的人。范围放宽是因为 2000 个样本本来就有波动。"""
    from copilot.api.routes.chat import in_agent_bucket

    _settings(monkeypatch, agent_enabled=False, agent_rollout=0.1)
    users = [uuid.uuid4() for _ in range(2000)]
    hit = sum(in_agent_bucket(u) for u in users)
    assert 120 < hit < 280, f"10% 的桶命中了 {hit}/2000"


def test_agent_enabled_overrides_rollout(monkeypatch):
    """总开关是给评测和本机验证用的，它要压过灰度比例。"""
    from copilot.api.routes.chat import in_agent_bucket

    _settings(monkeypatch, agent_enabled=True, agent_rollout=0.0)
    assert in_agent_bucket(uuid.uuid4())


async def test_bucketed_user_takes_the_agent_for_a_plain_question(maker, two_users, monkeypatch):
    """桶里的人，普通问答也走 Agent——这才是「全 Agent 化」。

    桶外的人不变：普通问答仍然走直路（M7 那套关键词路由）。
    """
    from copilot.api.routes.chat import _use_agent

    (alice_id, _), _, _ = two_users
    # `email` 不能省：M11 P4 之后 `_use_agent` 第一件事就是查白名单。
    # 给一个绝不会在 `AGENT_ALLOW_EMAILS` 里的地址，这样这道题量的仍然是**分桶**
    user = SimpleNamespace(id=alice_id, email="not-on-the-allowlist@test.local")
    async with maker() as s:
        _settings(monkeypatch, agent_enabled=False, agent_rollout=1.0)
        assert await _use_agent(s, user, "电子面单怎么设置", None) is True
        _settings(monkeypatch, agent_rollout=0.0)
        assert await _use_agent(s, user, "电子面单怎么设置", None) is False


# ---------- M10 P4：whoami / my_documents ----------


async def test_whoami_reuses_the_one_capability_text(maker, two_users):
    """⭐ 自我介绍只能有一份。

    寒暄短路那条路（`_canned_stream`）和这个工具必须给出同一段话——
    各写一份的话，加了个新能力改了一处忘了另一处，用户会看到
    「同一个助手对自己的两种说法」，而且没有任何报错。
    """
    from copilot.agent.tools import whoami
    from copilot.qa import canned_reply, small_talk_reply

    (alice_id, _), _, _ = two_users
    said: list[str] = []
    async with maker() as s:
        deps = _deps(s, alice_id)

        async def emit(kind, payload):
            if kind == "text":
                said.append(str(payload))

        deps.emit = emit
        back = await whoami(ctx(deps))

    assert "".join(said) == canned_reply("capability") == small_talk_reply("你能做什么")
    # 终结工具：正文已经直通给用户了，Agent 不该再复述
    assert deps.final_answer == "".join(said)
    assert "不要复述" in back


async def test_my_documents_only_lists_my_own(maker, two_users):
    """⭐ 隔离红线的第三处。公共库那 746 篇不属于任何人，也不该出现在这里。"""
    from copilot.agent.tools import my_documents

    (alice_id, bob_id), _, tag = two_users
    async with maker() as s:
        alice_sees = await my_documents(ctx(_deps(s, alice_id)))
        bob_sees = await my_documents(ctx(_deps(s, bob_id)))

    assert alice_id.hex[:6] in alice_sees
    assert bob_id.hex[:6] not in alice_sees, f"泄漏了！爱丽丝看到了鲍勃的文档：{alice_sees}"
    assert bob_id.hex[:6] in bob_sees
    assert alice_id.hex[:6] not in bob_sees


async def test_my_documents_says_so_when_empty(maker, two_users):
    """空的时候要明说「是空的」，别返回空串——空串会让模型以为工具坏了。"""
    from copilot.agent.tools import my_documents

    async with maker() as s:
        text = await my_documents(ctx(_deps(s, uuid.uuid4())))
    assert "空" in text
